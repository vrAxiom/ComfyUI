import os
import pandas as pd
from typing import Dict, Any
from openpyxl import load_workbook
from openpyxl.styles import Font

COLUMNS = [
    "Job Title",
    "Employer",
    "Applicant Name",
    "Current Designation",
    "Current Company",
    "Experience (Years)",
    "Experience (Months)",
    "Current CTC (LPA)",
    "Expected CTC (LPA)",
    "Current Location",
    "Preferred Locations",
    "Past Company",
    "Notice Period (Months)",
    "Education",
    "University",
    "Key Skills",
    "Date of Birth",
    "Mobile",
    "Email",
    "Offer In Hand (LPA)",
    "Ingested At",
    "Email Subject",
    "From Email",
    "AI Provider",
    "AI Model",
    "Confidence Score",
    "Response Link",
    "Contact Details Link",
    "Job Applicants Count",
    "Job Posted (Days)"
]

COLUMN_WIDTHS = {
    "Job Title": 30,
    "Employer": 25,
    "Applicant Name": 25,
    "Current Designation": 30,
    "Current Company": 25,
    "Experience (Years)": 18,
    "Experience (Months)": 18,
    "Current CTC (LPA)": 18,
    "Expected CTC (LPA)": 18,
    "Current Location": 20,
    "Preferred Locations": 35,
    "Past Company": 25,
    "Notice Period (Months)": 22,
    "Education": 40,
    "University": 30,
    "Key Skills": 60,
    "Date of Birth": 15,
    "Mobile": 18,
    "Email": 30,
    "Offer In Hand (LPA)": 18,
    "Ingested At": 25,
    "Email Subject": 50,
    "From Email": 35,
    "AI Provider": 14,
    "AI Model": 22,
    "Confidence Score": 18,
    "Response Link": 40,
    "Contact Details Link": 40,
    "Job Applicants Count": 20,
    "Job Posted (Days)": 20
}


def _ensure_parent_dir(path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)


def _format_sheet(filename: str, sheet_name: str):
    wb = load_workbook(filename)
    ws = wb[sheet_name]
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col, width in COLUMN_WIDTHS.items():
        # Map header names to column letters via header row
        for cell in ws[1]:
            if cell.value == col:
                ws.column_dimensions[cell.column_letter].width = width
                break
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(filename)


def _to_row(data: Dict[str, Any]) -> Dict[str, Any]:
    # Convert arrays to CSV strings
    def csv(v):
        if v is None:
            return None
        if isinstance(v, list):
            return ",".join(str(x) for x in v)
        return v

    return {
        "Job Title": data.get("job_title"),
        "Employer": data.get("employer"),
        "Applicant Name": data.get("applicant_name"),
        "Current Designation": data.get("current_designation"),
        "Current Company": data.get("current_company"),
        "Experience (Years)": data.get("experience_years"),
        "Experience (Months)": data.get("experience_months"),
        "Current CTC (LPA)": data.get("current_ctc_lpa"),
        "Expected CTC (LPA)": data.get("expected_ctc_lpa"),
        "Current Location": data.get("location_current"),
        "Preferred Locations": csv(data.get("location_preferred")),
        "Past Company": data.get("past_company"),
        "Notice Period (Months)": data.get("notice_period_months"),
        "Education": data.get("education"),
        "University": data.get("university"),
        "Key Skills": csv(data.get("key_skills")),
        "Date of Birth": data.get("date_of_birth"),
        "Mobile": data.get("mobile"),
        "Email": data.get("email"),
        "Offer In Hand (LPA)": data.get("offer_in_hand_lpa"),
        "Ingested At": data.get("ingested_at"),
        "Email Subject": data.get("email_subject"),
        "From Email": data.get("from_email"),
        "AI Provider": data.get("ai_provider"),
        "AI Model": data.get("ai_model"),
        "Confidence Score": data.get("confidence_score"),
        "Response Link": data.get("response_link"),
        "Contact Details Link": data.get("contact_details_link"),
        "Job Applicants Count": data.get("job_applicants_count"),
        "Job Posted (Days)": data.get("job_posted_days"),
    }


def write_to_excel(excel_path: str, data: Dict[str, Any]):
    _ensure_parent_dir(excel_path)
    row = _to_row(data)
    df = pd.DataFrame([row], columns=COLUMNS)

    if not os.path.exists(excel_path):
        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, sheet_name="Applications", index=False)
        _format_sheet(excel_path, "Applications")
        return

    # Append
    from openpyxl import load_workbook as _lw
    wb = _lw(excel_path)
    ws = wb["Applications"] if "Applications" in wb.sheetnames else wb.active
    start_row = ws.max_row + 1
    wb.close()

    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        df.to_excel(writer, sheet_name="Applications", index=False, header=False, startrow=start_row-1)
